#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @File  : sjl_baseio.py
# @Author: Shang
# @Date  : 2020/7/5


import codecs
from sklearn.model_selection import KFold
import numpy as np
import xlsxwriter
import csv
import xlrd
import pandas as pd
from xlutils.copy import copy
import openpyxl
import sys
csv.field_size_limit(sys.maxsize)



# 读取txt文件，清洗字符，去换行符
def readtxt(path):
    with open(path,'r',encoding='utf8') as f:
        lines=f.readlines()
        # 正则匹配
        lines=[i.strip().replace('\u3000','').replace('\n','').replace('\xa0','').replace('\ufeff','') for i in lines]
    return lines


# 读取txt文件
def readtxt_all(path):
    with open(path,'r',encoding='utf8') as f:
        lines=f.readlines()
    return lines


# 列表写入txt文件，一个元素为一行，有换行
def writetxt(path,data):
    with open(path,'w',encoding='utf8') as f:
        for i in data:
            f.write(str(i)+'\n')
    f.close()


# 字典直接写入txt文件
def write(path,data):
    with open(path,'w',encoding='utf8') as f:
        f.write(str(data))
    f.close()


# 列表写入txt文件，序列标注换行
def writetxt_line(path,data):
    with open(path,'w',encoding='utf8') as f:
        for i in data:
            if i[0] in ['？','。','！']:
                f.write(str(i)+'\n')
            else:
                f.write(str(i))
    f.close()


# 读取excel某列
def read_excel_col(path,sheet,col):
    # 打开文件
    workbook = xlrd.open_workbook(path)
    # 获取所有sheet
    # print(workbook.sheet_names()) # [u'sheet1', u'sheet2']
    # 获取sheet1
    sheet = workbook.sheets()[sheet]
    # 根据sheet索引或者名称获取sheet内容
    # sheet1 = workbook.sheet_by_name('Sheet1')
    # sheet的名称，行数，列数
    # sheet1.name, sheet1.nrows, sheet2.ncols
    # rows = sheet.row_values(3)  # 获取第四行内容
    cols = sheet.col_values(col)  # 获取第1列内容
    return cols


# 读取excel行
def read_excel_row(path,sheet):
    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheets()[sheet]
    rows=[]
    for i in range(sheet.nrows):
        rows.append(list(sheet.row_values(i)))  # 获取第四行内容
    return rows


# pd读取excel
def read2excel_pd(path):
    pd.read_excel(path, 'Sheet1', header=None)


# 将词典写入excel
def write_excel(path,dic):
    workbook = xlsxwriter.Workbook(path)
    worksheet = workbook.add_worksheet('sheet1')
    worksheet.write(0, 0, '词性')
    worksheet.write(0, 1, '频次')
    for k, i in zip(range(len(dic)), dic.keys()):
        j = dic[i]
        #print(i, j)
        worksheet.write(k+1, 0, i)
        worksheet.write(k+1, 1, j)
    workbook.close()


def write_excel_append(path,new_col_name,values):
    data = openpyxl.load_workbook(path)
    # print(data.sheetnames) # 输出所有工作页的名称
    # 取第一张表
    sheetnames = data.sheetnames
    table = data[sheetnames[0]]
    table = data.active
    # print(table.title) # 输出表名
    nrows = table.max_row  # 获得行数
    ncolumns = table.max_column  # 获得列数
    table.cell(1, ncolumns + 1).value = new_col_name
    for id, value in enumerate(values):
        # print(id, value)
        table.cell(id + 2, ncolumns + 1).value = str(value)
    data.save(path)


# 读取csv某行列
def read_csv(path,col_name):
    with open(path,'r',encoding='utf8',errors='ignore') as csvfile:
        reader = csv.DictReader(_.replace('\xa0','').replace('\u3000','').replace('\x00','').replace('\ufeff','') for _ in csvfile)
        column = [row[col_name] for row in reader]
    return column


# 列表去空字符元素
def strip_none(line):
    line = list(filter(None, line))
    return line


# CRF 自动分词
def deal(data):
    outdata = codecs.open('/home/ugrad/Shang/crf/crf词性标注/diease/alldata.txt', 'w', 'utf-8')
    # data=['中小学生/n 下周/t 就要/d 考试/vi 了/y ，/wd 许多/m 家长/n 孩子/n 鼓足/v 了/ule 劲/n ，/wd 结果/n 却/d 被/pbei 不期而遇/vl 的/ude1 流感/n 击倒/v ，/wd 不得不/d 请假/vi 挤/v 医院/n 。/wj 但/c 不/d 想/v 未/d 考/v 就/d 输/v ，/wd 有的/rz 孩子/n 去/vf 不/d 了/y 学校/n ，/wd 就/d 坚持/v 上/vf 校外/s 培训班/n ，/wd 结果/n 病毒/n 传播/v 开/v 来/vf ，/wd 更/d 多/a 孩子/n 中招/v ；/wf 着急/a 的/ude1 孩子/n 发/v 着/uzhe 高烧/n 想/v 着/uzhe 功课/n ，/wd 口中/s 还/d 念念有词/vl ……/ws 临/v 考/v 前/f 的/ude1 病/a 娃/ng 和/cc 家长/n 都/d 好/a 糟心/a 。/wj']
    cut_list=['。', '？', '！']
    for line in data:
        line=line.split(' ')
        line=filter(None,line)
        for word in line:
            word=word.split('/')
            # print(word)
            if len(word[0])==1:
                if word[0] in cut_list:
                    outdata.write(word[0] + '\tS-' + word[1] + '\n'+'\n')
                else:
                    outdata.write(word[0] + '\tS-' + word[1] + '\n')
            else:
                for i in range(len(word[0])):
                    if i == 0:
                        f = "B-"+word[1]
                    elif i == len(word[0]) - 1:
                        f = "E-"+word[1]
                    else:
                        f = "I-"+word[1]
                    outdata.write(word[0][i] + '\t' + f + '\r\n')


def kfold(data):
    KF = KFold(n_splits=10, shuffle=False, random_state=100)
    i=0
    for train_index, test_index in KF.split(data):
        # print("train_index:{},test_index:{}".format(train_index, test_index))
        data_train=data[list(train_index)[0]:list(train_index)[-1]]
        data_test=data[list(test_index)[0]:list(test_index)[-1]]
        path='./train'+str(i)+'.txt'
        path_test='./test'+str(i)+'.txt'
        # print(path)
        writetxt(path,data_train)
        writetxt(path_test,data_test)
        i+=1


def output_seq(file_name, sequences):
    with open(file_name, "w", encoding="utf-8", newline="\n") as fw:
        for seq in sequences:
            fw.write(seq)
            fw.write("\n")


def kflod2(sequence):
    k_10 = KFold(n_splits=10, shuffle=True, random_state=42)
    for fold_index, (train_index, test_index) in enumerate(k_10.split(sequence)):
        seqs_train_tag = np.array(sequence)[train_index].tolist()
        seqs_test_tag = np.array(sequence)[test_index].tolist()
        seqs_train = seqs_train_tag
        seqs_test = seqs_test_tag
        # print(seqs_train)
        # print(seqs_test)
        # output_seq("train{}.txt".format(fold_index), seqs_train)
        # output_seq("test{}.txt".format(fold_index), seqs_test)
